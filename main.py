import argparse
import logging
import re
from datetime import datetime, timedelta, timezone
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 常量定义
DEFAULT_INPUT_PATH = Path.cwd() / "(日志)表格视图.xlsx"
DEFAULT_OUTPUT_PATH = Path.cwd() / "tqsl.adi"
LOCAL_TIMEZONE = timezone(timedelta(hours=8))

# 字段映射
FIELD_MAPPING = {
    'CALL': '呼号',
    'MODE': '模式',
    'RST_RCVD': '对方信号',
    'RST_SENT': '己方信号',
    'RIG': '对方设备',
    'ANTENNA': '对方天馈',
    'RX_PWR': '对方功率',
    'QTH': '对方QTH',
    'MY_RIG': '己方设备',
    'MY_ANTENNA': '己方天馈',
    'TX_PWR': '己方功率',
    'MY_CITY': '己方QTH',
    'COMMENTS': '补充'
}

# 频率正则表达式
FREQUENCY_PATTERN = re.compile(r"(\d+\.\d+)(-\d+)?")


def write_adif_field(output: StringIO, field: str, data) -> None:
    """写入ADIF格式的字段"""
    if data is not None and data != "":
        s = str(data).strip()
        if '\n' in s:
            raise ValueError(f"{FIELD_MAPPING.get(field, field)}包含换行符")
        output.write(f"<{field.upper()}:{len(s)}>{s}\n")


def process_single_record(row, header_map) -> str:
    """处理单条记录，返回ADIF格式字符串，异常时返回None"""
    record_output = StringIO()

    try:
        # 获取时间字段索引
        time_col_idx = header_map.get('时间')
        time_value = row[time_col_idx].value
        if time_value is None:
            raise ValueError("缺少时间信息")
        # 设置时区为东八区
        local_dt = time_value.replace(tzinfo=LOCAL_TIMEZONE)
        # 转换为UTC时间
        utc_dt = local_dt.astimezone(timezone.utc)

        write_adif_field(record_output, 'QSO_DATE', utc_dt.strftime('%Y%m%d'))
        write_adif_field(record_output, 'TIME_ON', utc_dt.strftime('%H%M'))

        # 处理频率字段
        freq_col_idx = header_map.get('频率')
        freq_value = row[freq_col_idx].value
        if freq_value is None:
            raise ValueError("缺少频率信息")
        freq_str = str(freq_value)
        if match := FREQUENCY_PATTERN.match(freq_str):
            primary = float(match.group(1))
            if offset := match.group(2):
                write_adif_field(record_output, 'FREQ', primary + float(offset))
                write_adif_field(record_output, 'FREQ_RX', primary)
            else:
                write_adif_field(record_output, 'FREQ', primary)
        else:
            raise ValueError(f"频率格式不正确: {freq_str}")

        # 写入映射字段
        for adif_field, col_name in FIELD_MAPPING.items():
            col_idx = header_map.get(col_name)
            if col_idx is not None:
                value = row[col_idx].value
                write_adif_field(record_output, adif_field, value)

        record_output.write("<EOR>\n\n")
        return record_output.getvalue()

    except Exception as e:
        raise e
    finally:
        record_output.close()


def generate_adif(worksheet) -> str:
    """生成ADIF格式的日志内容，跳过异常记录"""
    logger.info("开始生成ADIF内容")
    output = StringIO()

    # 写入ADIF头部
    output.write("ADIFConvert by BH5UQJ\n")
    output.write("https://github.com/lingkai5wu/ADIFConvert\n")
    output.write("<ADIF_VER:5>3.1.6\n")
    output.write(f"<CREATED_TIMESTAMP:14>{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}\n")

    # 获取表头并创建映射
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    header_map = {cell.value: idx for idx, cell in enumerate(header_row) if cell.value is not None}

    # 验证必要的列是否存在
    required_columns = ['呼号', '时间', '频率']
    for col in required_columns:
        if col not in header_map:
            raise ValueError(f"Excel文件缺少必要的列: {col}")

    # 处理数据行
    total_rows = worksheet.max_row - 1  # 减去表头行
    success_count = 0
    error_count = 0

    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
        try:
            record = process_single_record(row, header_map)
            output.write(record)
            success_count += 1
        except Exception as e:
            logger.error(f"处理第 {row_num} 行时出错，已跳过: {e}")
            error_count += 1
            continue

        # 进度日志
        if (row_num - 1) % 100 == 0 or row_num - 1 == total_rows:
            logger.info(f"已处理 {row_num - 1}/{total_rows} 条记录，成功: {success_count}, 失败: {error_count}")

    logger.info(f"ADIF内容生成完成。总记录: {total_rows}, 成功写入: {success_count}, 跳过异常记录: {error_count}")
    return output.getvalue()


def main():
    logger.info(f"输入文件: {input_path}, 输出文件: {output_path}")

    try:
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")

        # 读取Excel并生成ADIF
        logger.info("开始读取Excel文件")
        workbook = load_workbook(filename=str(input_path), data_only=True)
        # 使用第一个工作表
        worksheet = workbook.active
        logger.info(f"成功读取Excel文件，共 {worksheet.max_row - 1} 条记录")

        adif_content = generate_adif(worksheet)

        # 写入输出文件
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(adif_content, encoding='utf-8')
        logger.info(f"成功生成ADIF文件: {output_path}")

    except Exception as e:
        logger.error(e, exc_info=True)
        raise SystemExit(1)


def get_user_confirmation(prompt, message=None, details=None, default_confirm=False):
    """ 获取用户确认的通用函数"""
    # 打印主要信息
    if message:
        print(message)
    # 如果有详细信息，逐条打印
    if details:
        for item in details:
            print(item)
    # 根据默认选择设置提示信息
    prompt += "[Y/n]" if default_confirm else "[y/N]"

    while True:
        response = input(prompt).strip().lower()
        # 处理空输入（使用默认选择）
        if response == '':
            return default_confirm
        # 确认选项
        if response in ['y', 'yes', '是']:
            return True
        # 拒绝选项
        if response in ['n', 'no', '否']:
            return False
        # 无效输入
        print("输入无效，请重新输入")


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='将Excel日志转换为ADIF格式')
    parser.add_argument('-i', '--input', type=str, help=f'输入Excel文件路径，默认: {DEFAULT_INPUT_PATH}')
    parser.add_argument('-o', '--output', type=str, help=f'输出ADIF文件路径，默认: {DEFAULT_OUTPUT_PATH}')
    args = parser.parse_args()

    # 确定输入输出路径
    input_path = Path(args.input) if args.input else DEFAULT_INPUT_PATH
    output_path = Path(args.output) if args.output else DEFAULT_OUTPUT_PATH

    while True:
        main()
        if not get_user_confirmation('是否再次运行？'):
            break

# pyinstaller -F -n "ADIFConvert_V0.2" .\main.py
